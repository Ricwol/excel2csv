import csv
from pathlib import Path

from openpyxl import load_workbook


def xlsx_to_csv(excel_file: Path, csv_file: Path) -> None:
    """Convert Excel file to csv file."""
    workbook = load_workbook(excel_file)
    sheet = workbook.active

    with csv_file.open('w', newline='') as writer:
        csv_writer = csv.writer(writer)
        csv_writer.writerows(sheet.iter_rows(values_only=True))
